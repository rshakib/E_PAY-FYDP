import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'e_banking', 'backend')))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
WRAP = Alignment(wrap_text=True, vertical='top')

COLUMNS = [
    ("Test ID", 10),
    ("Module", 12),
    ("Test Name", 25),
    ("Description", 35),
    ("Preconditions", 25),
    ("Step #", 6),
    ("Step Description", 40),
    ("Input Data", 25),
    ("Expected Result", 30),
    ("Actual Result", 30),
    ("Pass/Fail", 8),
    ("Screenshot Ref", 25),
    ("Notes / Bugs", 30),
]


def setup_sheet(ws, title):
    ws.title = title
    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M1"


def add_row(ws, row_num, data):
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.alignment = WRAP
        cell.border = BORDER
        if col_idx == 11:  # Pass/Fail
            if value == "PASS":
                cell.fill = PASS_FILL
            elif value == "FAIL":
                cell.fill = FAIL_FILL


# ═══════════════════════════════════════════════════════════
# TEST 1: Unit Testing - Cryptographic Functions
# ═══════════════════════════════════════════════════════════
ws1 = wb.active
setup_sheet(ws1, "Test 1 - Unit Crypto")

t1_cases = [
    # TC-1.1
    ["TC-1.1", "Crypto", "AES-256 Encrypt → Decrypt Roundtrip",
     "Verify that AES-256 CBC encryption followed by decryption returns the original plaintext message and HMAC.",
     "CryptoEngine class imported; known values for K2, Bp, T available", 1,
     "Initialize CryptoEngine with a sample message M = 'Receiver:Alice|Amt:500'",
     "M = 'Receiver:Alice|Amt:500'; K1 = 'key_k1'", "generate_hmac(K1, M) returns a 64-char hex F1", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Call encrypt_data(M, F1, K2, Bp, T)", "K2 = sample_256bit_hex; Bp = '123456'; T = ISO_timestamp",
     "Returns {'payload': base64_str, 'iv': base64_str}", "", "", "", ""],
    ["", "", "", "", "", 3,
     "Call decrypt_data(payload, iv, K2, Bp, T)", "Same K2, Bp, T as encryption",
     "Returns {'M': original_message, 'F1': original_hmac}", "", "", "", ""],
    ["", "", "", "", "", 4,
     "Assert decrypted M equals original M and F1 matches", "—",
     "Assertions pass; decrypted M == 'Receiver:Alice|Amt:500'", "", "", "", ""],

    # TC-1.2
    ["TC-1.2", "Crypto", "HMAC-SHA256 Deterministic Output",
     "Verify that HMAC generation produces identical output for identical inputs, and different output for different inputs.",
     "CryptoEngine initialized", 1,
     "Call generate_hmac('test_hmac_key', 'Receiver:Bob|Amt:1000') twice",
     "key='test_hmac_key', msg='Receiver:Bob|Amt:1000'",
     "Both calls return identical 64-char hex string", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Call generate_hmac with a different key: 'test_hmac_key_x'",
     "key='test_hmac_key_x', same message", "Output differs from first hash", "", "", "", ""],
    ["", "", "", "", "", 3,
     "Assert hash1 == hash2 and hash1 != hash3",
     "—", "Assertions pass", "", "", "", ""],

    # TC-1.3
    ["TC-1.3", "Crypto", "Key Derivation Deterministic (K2+Bp+T)",
     "Verify that the internal _derive_aes_key method produces the same 32-byte key for identical K2, Bp, T inputs, and different keys when inputs change.",
     "CryptoEngine initialized; sample K2, Bp, T known", 1,
     "Call _derive_aes_key(K2, Bp, T) twice",
     "Same K2, Bp, T values", "Both calls return identical 32-byte key", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Call with different T (append 'x')",
     "T2 = T + 'x'", "Derived key differs from step 1", "", "", "", ""],
    ["", "", "", "", "", 3,
     "Call with different K2 (append 'x')",
     "K2' = K2 + 'x'", "Derived key differs from step 1", "", "", "", ""],
    ["", "", "", "", "", 4,
     "Assert length of derived key == 32 bytes",
     "—", "len(key) == 32", "", "", "", ""],

    # TC-1.4
    ["TC-1.4", "Crypto", "Altered Ciphertext → Decryption Fails",
     "Verify that modifying even one byte of the ciphertext causes decrypt_data to return None (integrity protection).",
     "A valid ciphertext payload exists from a prior encryption", 1,
     "Decode payload from base64, flip one byte at position 5, re-encode",
     "Original valid payload + IV", "Corrupted ciphertext produced", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Call decrypt_data(corrupted_payload, iv, K2, Bp, T)",
     "Same IV, K2, Bp, T as original encryption",
     "Returns None", "", "", "", ""],
    ["", "", "", "", "", 3,
     "Assert decrypted is None", "—", "Assertion passes", "", "", "", ""],

    # TC-1.5
    ["TC-1.5", "Crypto", "Altered HMAC → Mismatch Detection",
     "Verify that if F1 in decrypted data is tampered with, F2 (recomputed by server) differs, making the attack detectable.",
     "CryptoEngine initialized; known K1 and message", 1,
     "Generate original F1 = HMAC(K1, M) and tampered F1' = HMAC(K1_tampered, M)",
     "K1 = 'hmac_key_for_test'; K1_tampered = 'hmac_key_for_test_tampered'; M = 'Receiver:Dave|Amt:300'",
     "F1 and F1' are different hex strings", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Recompute F2 = HMAC(K1, M)", "Original correct K1", "F2 == F1 (original)", "", "", "", ""],
    ["", "", "", "", "", 3,
     "Assert F1' != F2", "—", "Mismatch confirmed; tampered HMAC would fail server-side 403 check", "", "", "", ""],

    # TC-1.6
    ["TC-1.6", "Crypto", "Same Plaintext + Different T → Different Ciphertext",
     "Verify that encrypting the same message with different T values produces different ciphertexts (timestamp-binding).",
     "CryptoEngine initialized; message and F1 fixed", 1,
     "Encrypt with T1 = '2026-01-01T00:00:00Z'",
     "Message = 'Receiver:Eve|Amt:400'; F1 = HMAC(K1, msg)", "Ciphertext payload_1 and IV_1", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Encrypt same message with T2 = '2026-06-01T00:00:00Z'",
     "Same Message, same F1, K2, Bp", "Ciphertext payload_2 and IV_2", "", "", "", ""],
    ["", "", "", "", "", 3,
     "Assert payload_1 != payload_2 and IV_1 != IV_2",
     "—", "Both payload and IV differ; timestamp binding confirmed", "", "", "", ""],

    # TC-1.7
    ["TC-1.7", "Crypto", "PBKDF2 Password Stretching",
     "Verify that stretch_password produces consistent 64-char hex output for same inputs, and different output for different NID.",
     "CryptoEngine initialized", 1,
     "Call stretch_password('securePass123!', 'NID12345') twice",
     "password='securePass123!', nid='NID12345'",
     "Both calls produce identical 64-char hex string", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Call with same password but different NID 'NID99999'",
     "password='securePass123!', nid='NID99999'",
     "Result differs from step 1", "", "", "", ""],
    ["", "", "", "", "", 3,
     "Assert output length == 64", "—", "len == 64 (32 bytes hex-encoded)", "", "", "", ""],

    # TC-1.8
    ["TC-1.8", "Crypto", "Multiple Encrypt/Decrypt Cycles",
     "Verify that multiple rounds of encrypt/decrypt with various amounts all succeed (50, 9999.99, 0.01).",
     "CryptoEngine initialized", 1,
     "Encrypt and decrypt messages with amounts 50, 9999.99, 0.01",
     "Messages with receiver='Frank|Grace|Henry' and varying amounts",
     "All 3 roundtrips return original messages", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Assert each decrypted M matches original", "—",
     "Assertions pass for all 3 test cases", "", "", "", ""],
]

row = 2
for case in t1_cases:
    add_row(ws1, row, case)
    row += 1

# ═══════════════════════════════════════════════════════════
# TEST 2: Integration Testing - Transaction Flow
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet()
setup_sheet(ws2, "Test 2 - Integration")

t2_cases = [
    # TC-2.0
    ["TC-2.0", "Integration", "Backend Health Check",
     "Verify the Flask backend is reachable and responds with 200 OK at /health.",
     "Flask server running on localhost:5001", 1,
     "Send GET /health request", "No input required",
     "HTTP 200; JSON {'status': 'ok', 'message': 'E-Banking API is running'}", "", "", "", ""],

    # TC-2.1
    ["TC-2.1", "Integration", "User Registration",
     "Verify that a new user can register via POST /register with username, password, NID, and activation code.",
     "Backend running; unique username available", 1,
     "Send POST /register with new user credentials",
     "username='testuser_XXX', password='Test@12345', nid='NID998877', activationCode='ACT5566'",
     "HTTP 201; JSON {'status': 'success', 'message': 'Account created successfully'}", "", "", "", ""],

    # TC-2.2
    ["TC-2.2", "Integration", "User Login Returns Crypto Keys",
     "Verify that POST /login returns K1, K2, Bp, T, balance, accountId on successful authentication.",
     "User already registered from TC-2.1", 1,
     "Send POST /login with username and password",
     "username='testuser_XXX', password='Test@12345'",
     "HTTP 200; JSON with fields: token, user.{id, username, k1, k2, bp, t, balance, accountId}", "", "", "", ""],

    # TC-2.3
    ["TC-2.3", "Integration", "Valid Transfer (Success)",
     "Verify that a properly encrypted transfer with correct HMAC returns HTTP 200 and updates T and balance.",
     "Logged in with valid token; crypto keys (K1, K2, Bp, T) known", 1,
     "Encrypt 'Receiver:testuser_XXX|Amt:100' with server's K1, K2, Bp, T",
     "message='Receiver:testuser_XXX|Amt:100'; F1=HMAC(K1, msg); encrypt(K2,Bp,T)",
     "Encrypted payload + IV generated", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Send POST /transfer with encrypted payload",
     "payload and IV from step 1; Bearer token in header",
     "HTTP 200; JSON with status='success', new_t, new_balance", "", "", "", ""],

    # TC-2.4
    ["TC-2.4", "Integration", "Tampered HMAC → 403 Rejection",
     "Verify that a transfer with incorrect HMAC (F1 != F2) is rejected with HTTP 403.",
     "Logged in with valid token; crypto keys known", 1,
     "Encrypt with wrong K1 for HMAC generation",
     "F1=HMAC('wrong_key', 'Receiver:Fake|Amt:999'); encrypt with correct K2,Bp,T",
     "Encrypted payload with invalid F1 generated", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Send POST /transfer with invalid-HMAC payload",
     "Tampered payload + valid IV + Bearer token",
     "HTTP 403; JSON with status='error', message='Data integrity compromised (HMAC mismatch)'", "", "", "", ""],

    # TC-2.5
    ["TC-2.5", "Integration", "Transfer to Non-Existent Receiver → 404",
     "Verify that transferring to a username that doesn't exist returns HTTP 404.",
     "Logged in; valid T from TC-2.3 result", 1,
     "Encrypt 'Receiver:NonexistentUser999|Amt:50'",
     "message='Receiver:NonexistentUser999|Amt:50'; use updated T from previous transfer",
     "Encrypted payload generated", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Send POST /transfer to nonexistent receiver",
     "Payload from step 1 + Bearer token",
     "HTTP 404; JSON with status='error', message='Receiver not found'", "", "", "", ""],

    # TC-2.6
    ["TC-2.6", "Integration", "Insufficient Balance → 400 'futile'",
     "Verify that transferring an amount greater than the current balance returns HTTP 400 with status='futile'.",
     "Logged in; balance known (~4900 after TC-2.3 deduction)", 1,
     "Encrypt message with amount >> current balance (e.g., 999999)",
     "message='Receiver:testuser_XXX|Amt:999999'; use current T",
     "Encrypted payload generated", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Send POST /transfer with excessive amount",
     "Payload from step 1 + Bearer token",
     "HTTP 400; JSON with status='futile', message='Insufficient balance'", "", "", "", ""],

    # TC-2.7
    ["TC-2.7", "Integration", "Balance Update Verification",
     "Verify sender's balance decreased correctly after the successful transfer in TC-2.3 (5000 - 100 = 4900).",
     "Logged in; TC-2.3 transfer completed", 1,
     "Send GET /user/{username}", "Bearer token in header",
     "HTTP 200; JSON user.balance = 4900.0 (within floating-point tolerance)", "", "", "", ""],

    # TC-2.8
    ["TC-2.8", "Integration", "Transaction in History",
     "Verify that the successful transfer from TC-2.3 appears in the transaction history with status='success'.",
     "Logged in; at least one transfer completed", 1,
     "Send GET /transactions/{username}", "Bearer token in header",
     "HTTP 200; JSON with transactions array; at least 1 entry; latest has status='success'", "", "", "", ""],

    # TC-2.9
    ["TC-2.9", "Integration", "Frontend Login via UI",
     "Verify that the login page loads and user can log in through the browser UI.",
     "Frontend served at https://localhost:5001; test user exists", 1,
     "Navigate to /login page in browser",
     "URL: /login", "Login form is displayed with username/password fields", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Type test username and password and click Login",
     "username='testuser_XXX', password='Test@12345'",
     "Dashboard page loads; no error messages", "", "", "", ""],

    # TC-2.10
    ["TC-2.10", "Integration", "Frontend Serves Correctly",
     "Verify that the frontend homepage loads with correct title and basic layout.",
     "Frontend served at https://localhost:5001", 1,
     "Navigate to / in browser",
     "URL: https://localhost:5001/",
     "Page loads; title contains 'e-pay'; no console errors", "", "", "", ""],
]

row = 2
for case in t2_cases:
    add_row(ws2, row, case)
    row += 1

# ═══════════════════════════════════════════════════════════
# TEST 3: Security Testing - Replay Attack Prevention
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet()
setup_sheet(ws3, "Test 3 - Replay Security")

t3_cases = [
    # TC-3.0
    ["TC-3.0", "Replay", "Test User Setup",
     "Register and login a dedicated user for replay attack tests.",
     "Backend running", 1,
     "Register a new user and login",
     "username='replayusr_XXX', password='Replay@99'",
     "Login returns token + crypto keys (K1, K2, Bp, T)", "", "", "", ""],

    # TC-3.1
    ["TC-3.1", "Replay", "Same Payload Sent Twice → Second Rejected",
     "Verify that submitting the exact same encrypted payload twice results in the second attempt being rejected (due to T rotation or balance change).",
     "Logged in; valid T and crypto keys", 1,
     "Encrypt a valid transfer: 'Receiver:replayusr_XXX|Amt:50' and send it",
     "message='Receiver:replayusr_XXX|Amt:50'; F1=HMAC(K1, msg); encrypt(K2,Bp,T)",
     "HTTP 200; transfer succeeds, T is updated", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Resend the exact same payload and IV (no changes)",
     "Same payload, same IV, same Bearer token",
     "HTTP 400/403/401; second attempt rejected", "", "", "", ""],

    # TC-3.2
    ["TC-3.2", "Replay", "Stale/Old T Payload Rejected",
     "Verify that a payload encrypted with an old (previous) T value is rejected after T has been rotated by a successful transaction.",
     "T was updated after TC-3.1; old_T is known", 1,
     "Encrypt a new message using the OLD T value (before rotation)",
     "message='Receiver:replayusr_XXX|Amt:25'; use old_T; correct K1, K2, Bp",
     "Encrypted payload generated with stale T", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Send payload encrypted with old T",
     "Stale payload + current Bearer token",
     "HTTP 400/401/403; stale T payload rejected (key derivation mismatch)", "", "", "", ""],

    # TC-3.3
    ["TC-3.3", "Replay", "Corrupted Ciphertext → Rejected",
     "Verify that a transfer with a corrupted ciphertext (bit-flipped) is rejected by the server.",
     "Valid encrypted payload available", 1,
     "Decode payload from base64, flip one byte at position -3, re-encode",
     "Original valid payload", "Corrupted ciphertext produced", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Send corrupted payload with original IV and valid token",
     "Corrupted payload + valid IV + Bearer token",
     "HTTP 400/401/403; decryption fails", "", "", "", ""],

    # TC-3.4
    ["TC-3.4", "Replay", "Missing Required Fields → 400",
     "Verify that omitting payload or IV from the transfer request results in HTTP 400.",
     "Logged in", 1,
     "Send POST /transfer with only username (no payload)",
     "JSON: {'username': 'replayusr_XXX'} + Bearer token",
     "HTTP 400; JSON with error message about missing fields", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Send POST /transfer with payload but no IV",
     "JSON: {'username': 'replayusr_XXX', 'payload': 'abc123'} + Bearer token",
     "HTTP 400; JSON with error message about missing iv", "", "", "", ""],

    # TC-3.5
    ["TC-3.5", "Replay", "No Auth Token → 401",
     "Verify that a transfer request without an Authorization header is rejected with 401.",
     "Backend running", 1,
     "Send POST /transfer without Bearer token",
     "JSON: {'username': 'replayusr_XXX', 'payload': 'test', 'iv': 'test'} (no Authorization header)",
     "HTTP 401; JSON with status='error', message='Unauthorized'", "", "", "", ""],

    # TC-3.6
    ["TC-3.6", "Replay", "UI Screenshots Capture",
     "Capture screenshots of the login page and dashboard for the replay test user to document the test environment.",
     "Frontend served at BACKEND_URL; test user exists", 1,
     "Launch Playwright, navigate to /login page",
     "URL: https://localhost:5001/login",
     "Login page screenshot captured as 3.5_login_page.png", "", "", "", ""],
    ["", "", "", "", "", 2,
     "Fill in test user credentials and click Login",
     "username='repl_XXXXX', password='Replay@1'",
     "Dashboard loads; screenshot captured as 3.6_dashboard.png", "", "", "", ""],
]

row = 2
for case in t3_cases:
    add_row(ws3, row, case)
    row += 1

# ═══════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════
OUTPUT = os.path.join(os.path.dirname(__file__), "Test_Cases.xlsx")
wb.save(OUTPUT)
print(f"Spreadsheet created: {OUTPUT}")
print(f"  Test 1 - Unit Crypto: {ws1.max_row - 1} rows")
print(f"  Test 2 - Integration: {ws2.max_row - 1} rows")
print(f"  Test 3 - Replay Security: {ws3.max_row - 1} rows")
